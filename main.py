import os, re, io, json, base64, sqlite3, tempfile, shutil
from datetime import datetime, date
from contextlib import contextmanager
from fastapi import FastAPI, UploadFile, File, Query, HTTPException
from fastapi.responses import FileResponse, StreamingResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from dateutil import parser as dateparser

app = FastAPI()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

FIELDS = ["sn", "policyno", "name", "doc", "fup", "sumass", "plan", "mode", "premium", "mobileno", "status"]
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# ── DB (Turso cloud or local SQLite) ────────────────────────────────────────────

TURSO_URL = os.environ.get("TURSO_DATABASE_URL")
TURSO_TOKEN = os.environ.get("TURSO_AUTH_TOKEN")
USE_TURSO = bool(TURSO_URL and TURSO_TOKEN)

if USE_TURSO:
    import turso.sync as turso_sync
    DB_PATH = os.path.join(BASE_DIR, "local_replica.db")
else:
    DB_PATH = os.path.join(BASE_DIR, "master.db")

def dict_factory(cursor, row):
    """Row factory that works with both sqlite3 and Turso cursors."""
    fields = [col[0] for col in cursor.description]
    return dict(zip(fields, row))

def get_db():
    if USE_TURSO:
        conn = turso_sync.connect(
            DB_PATH,
            remote_url=TURSO_URL,
            auth_token=TURSO_TOKEN,
        )
    else:
        conn = sqlite3.connect(DB_PATH)
    conn.row_factory = dict_factory
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def db_push():
    """Push local changes to Turso cloud (no-op in local dev)."""
    if USE_TURSO:
        conn = get_db()
        try:
            conn.push()
        finally:
            conn.close()

def db_pull():
    """Pull remote changes from Turso cloud (no-op in local dev)."""
    if USE_TURSO:
        conn = get_db()
        try:
            conn.pull()
        finally:
            conn.close()

def init_db():
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS policies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            policyno TEXT UNIQUE NOT NULL,
            sn TEXT, name TEXT, doc TEXT, fup TEXT, sumass TEXT,
            plan TEXT, mode TEXT, premium TEXT, mobileno TEXT, status TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS upload_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            sheet_name TEXT,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            record_count INTEGER DEFAULT 0,
            records_json TEXT
        )""")
    db_push()

init_db()
# Pull latest data from cloud on startup
db_pull()

def save_upload_copy(filename: str, content: bytes):
    """Save a timestamped backup copy of every uploaded file."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    name, ext = os.path.splitext(filename)
    safe_name = re.sub(r'[^\w\-.]', '_', name)
    dest = os.path.join(UPLOAD_DIR, f"{safe_name}_{ts}{ext}")
    with open(dest, "wb") as f:
        f.write(content)
    return dest

# ── Normalization ───────────────────────────────────────────────────────────────

COL_MAP = {
    "sn": "sn", "sno": "sn", "srno": "sn", "serialno": "sn", "sr": "sn", "slno": "sn", "serial": "sn", "no": "sn",
    "policyno": "policyno", "policynumber": "policyno", "policynum": "policyno", "policy": "policyno", "polno": "policyno",
    "name": "name", "holdername": "name", "insuredname": "name", "clientname": "name",
    "policyname": "name", "policyholdername": "name", "insured": "name",
    "nameofassured": "name", "assuredname": "name", "nameassured": "name",
    "doc": "doc", "dateofcommencement": "doc", "commencementdate": "doc", "startdate": "doc",
    "dtofcomm": "doc", "dtofcommencement": "doc", "commdate": "doc", "dateofcommence": "doc", "dateofcomm": "doc",
    "fup": "fup", "firstunpaidpremium": "fup", "duedate": "fup", "nextdue": "fup",
    "nextduedate": "fup", "unpaidpremium": "fup", "firstunpaid": "fup",
    "sumass": "sumass", "sumassured": "sumass", "suminsured": "sumass", "sa": "sumass",
    "plan": "plan", "planname": "plan", "planno": "plan", "plantype": "plan",
    "mode": "mode", "premiummode": "mode", "paymentmode": "mode", "frequency": "mode", "paymode": "mode",
    "premium": "premium", "premiumpayable": "premium", "premiumamount": "premium",
    "amt": "premium", "amount": "premium", "prem": "premium",
    "mobileno": "mobileno", "mobile": "mobileno", "phone": "mobileno", "phoneno": "mobileno",
    "contactno": "mobileno", "contact": "mobileno", "cellno": "mobileno",
    "mobilenumber": "mobileno", "phonenumber": "mobileno", "cell": "mobileno", "mob": "mobileno",
    "status": "status", "policystatus": "status", "paymentstatus": "status",
}

# Substrings that should never match fuzzy column detection
_FUZZY_EXCLUDE = {"unnamed", "untitled", "column", "header", "field"}

def normalize_col(name):
    if not name: return None
    cleaned = re.sub(r"[^a-z0-9]", "", str(name).lower().strip())
    if cleaned in COL_MAP: return COL_MAP[cleaned]
    # skip fuzzy match for known non-column strings
    if any(ex in cleaned for ex in _FUZZY_EXCLUDE): return None
    # fuzzy substring matching for messier headers
    # require the key to cover a significant portion of the cleaned string
    # to avoid false positives on title rows like "licpolicyregisterform"
    for key, val in COL_MAP.items():
        if len(key) >= 4 and key in cleaned:
            if len(key) / len(cleaned) >= 0.4:
                return val
    return None

STATUS_MAP = {
    "autodebit": "autodebit", "auto debit": "autodebit",
    "dailycollection": "dailycollection", "daily collection": "dailycollection",
    "branchpaid": "branchpaid", "branch paid": "branchpaid", "branchpaidonly": "branchpaid",
    "paid": "paid", "due": "due",
}

def normalize_status(val):
    if not val or str(val).strip() == "": return None
    cleaned = re.sub(r"[^a-z ]", "", str(val).lower().strip()).strip()
    if cleaned in STATUS_MAP: return STATUS_MAP[cleaned]
    nospace = cleaned.replace(" ", "")
    if nospace in STATUS_MAP: return STATUS_MAP[nospace]
    return None

def clean_val(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    s = str(v).strip()
    return s if s else None

def normalize_policyno(v):
    s = clean_val(v)
    if not s: return None
    # strip commas, spaces, dots that Excel number formatting may add
    s = re.sub(r"[,\s\.]", "", s)
    return s.upper()

# ── Upsert ──────────────────────────────────────────────────────────────────────

def parse_date(s):
    if not s: return None
    try: return dateparser.parse(str(s), dayfirst=True)
    except: return None

def upsert_records(records):
    inserted, updated = 0, 0
    non_pno = [f for f in FIELDS if f != "policyno"]
    with get_db() as conn:
        for rec in records:
            pno = normalize_policyno(rec.get("policyno"))
            if not pno: continue
            # normalize status
            if rec.get("status"): rec["status"] = normalize_status(rec["status"]) or clean_val(rec["status"])
            existing = conn.execute("SELECT * FROM policies WHERE UPPER(TRIM(policyno))=?", (pno,)).fetchone()
            if existing:
                updates, params = [], []
                for f in non_pno:
                    new_val = clean_val(rec.get(f))
                    if not new_val: continue
                    old_val = existing[f]
                    if f == "fup":
                        new_dt, old_dt = parse_date(new_val), parse_date(old_val)
                        if new_dt and old_dt and new_dt > old_dt:
                            updates.append(f"{f}=?"); params.append(new_val)
                        elif not old_val:
                            updates.append(f"{f}=?"); params.append(new_val)
                    elif not old_val:
                        updates.append(f"{f}=?"); params.append(new_val)
                if updates:
                    updates.append("updated_at=?"); params.append(datetime.now().isoformat())
                    params.append(pno)
                    conn.execute(f"UPDATE policies SET {','.join(updates)} WHERE UPPER(TRIM(policyno))=?", params)
                    updated += 1
            else:
                vals = {f: clean_val(rec.get(f)) for f in FIELDS}
                vals["policyno"] = pno
                vals["updated_at"] = datetime.now().isoformat()
                cols = list(vals.keys())
                conn.execute(
                    f"INSERT INTO policies ({','.join(cols)}) VALUES ({','.join('?' for _ in cols)})",
                    [vals[c] for c in cols]
                )
                inserted += 1
    return inserted, updated

# ── File processors ─────────────────────────────────────────────────────────────

def _score_header_row(row_values):
    """Score how many cells in a row look like recognized column headers."""
    hits = 0
    seen = set()
    for v in row_values:
        norm = normalize_col(v)
        if norm and norm not in seen:
            hits += 1
            seen.add(norm)
    return hits

def _text_ratio(row_values):
    """Return fraction of non-empty cells that are text (non-numeric).
    Headers tend to be all-text; data rows tend to be mixed."""
    text_count, total = 0, 0
    for v in row_values:
        if v is None: continue
        s = str(v).strip()
        if not s: continue
        total += 1
        try:
            float(s.replace(",", "").replace(" ", ""))
        except ValueError:
            text_count += 1
    return text_count / total if total else 0

def find_header_row(sheet_bytes, engine, sheet_name, max_scan=15):
    """Scan the first `max_scan` rows to find the best header row.

    Returns the 0-based header index to pass to pd.read_excel(header=...).
    Falls back to 0 if no better row is found.
    Uses both column-name recognition and text-ratio heuristics.
    """
    df_raw = pd.read_excel(
        io.BytesIO(sheet_bytes), sheet_name=sheet_name,
        engine=engine, header=None, dtype=str,
        nrows=max_scan,
    )
    best_idx, best_score = 0, 0
    for idx in range(len(df_raw)):
        row_vals = [str(v) if pd.notna(v) else None for v in df_raw.iloc[idx]]
        non_empty = sum(1 for v in row_vals if v and str(v).strip())
        if non_empty == 0: continue
        col_hits = _score_header_row(row_vals)
        text_r = _text_ratio(row_vals)
        # Weighted score: column name matches are strongest signal,
        # high text ratio is secondary (headers are mostly text)
        score = col_hits * 10 + (text_r * non_empty * 2)
        if score > best_score:
            best_score = score
            best_idx = idx
    # Accept if we found at least 1 recognized column header
    return best_idx if best_score >= 10 else 0

def normalize_df(df):
    col_mapping = {}
    for c in df.columns:
        norm = normalize_col(c)
        if norm and norm not in col_mapping.values():
            col_mapping[c] = norm
    df = df.rename(columns=col_mapping)
    keep = [c for c in df.columns if c in FIELDS]
    if "policyno" not in keep: return []
    df = df[keep]
    df = df.dropna(subset=["policyno"])
    return df.to_dict(orient="records")

def _scrub_nans(records):
    """Replace float NaN values with None so JSON serialization works."""
    import math
    for rec in records:
        for k, v in rec.items():
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                rec[k] = None
            elif isinstance(v, str) and v.strip().lower() in ("nan", "inf", "-inf"):
                rec[k] = None
    return records

def _save_upload_batch(filename, sheet_data):
    """Save per-sheet upload records for history viewing."""
    if not sheet_data:
        return
    with get_db() as conn:
        for sd in sheet_data:
            clean = _scrub_nans(list(sd["records"]))
            conn.execute(
                "INSERT INTO upload_batches (filename, sheet_name, uploaded_at, record_count, records_json) VALUES (?,?,?,?,?)",
                (filename, sd["sheet_name"], datetime.now().isoformat(), len(clean), json.dumps(clean))
            )

def process_spreadsheet(content, filename):
    ext = filename.rsplit(".", 1)[-1].lower()
    all_records = []
    sheet_data = []
    if ext == "csv":
        df = pd.read_csv(io.BytesIO(content), dtype=str)
        records = normalize_df(df)
        all_records = records
        sheet_data.append({"sheet_name": filename, "records": records})
    else:
        engine = "xlrd" if ext == "xls" else "openpyxl"
        xls = pd.ExcelFile(io.BytesIO(content), engine=engine)
        for sheet in xls.sheet_names:
            header_idx = find_header_row(content, engine, sheet)
            df = pd.read_excel(xls, sheet_name=sheet, dtype=str, header=header_idx)
            records = normalize_df(df)
            all_records.extend(records)
            sheet_data.append({"sheet_name": sheet, "records": records})
    _save_upload_batch(filename, sheet_data)
    if not all_records: return 0, 0, "No recognizable policy data found"
    ins, upd = upsert_records(all_records)
    return ins, upd, None

def process_image(content, filename):
    from google import genai
    prompt = ("Extract all insurance policy table data from this image. Return ONLY a raw JSON array "
              "of objects with these exact keys: sn, policyno, name, doc, fup, sumass, plan, mode, "
              "premium, mobileno, status. Use null for missing fields. No markdown, no explanation, raw JSON only.")
    client = genai.Client()
    for attempt in range(2):
        try:
            resp = client.models.generate_content(
                model="gemini-3.5-flash",
                contents=[
                    genai.types.Part.from_bytes(data=content, mime_type="image/jpeg"),
                    prompt
                ]
            )
            text = resp.text.strip()
            if text.startswith("```"): text = re.sub(r"^```\w*\n?", "", text).rstrip("`").strip()
            records = json.loads(text)
            if not isinstance(records, list): return 0, 0, "OCR returned non-array JSON"
            ins, upd = upsert_records(records)
            return ins, upd, None
        except Exception as e:
            if attempt == 0: continue
            return 0, 0, f"OCR failed: {e}"


def process_docx(content, filename):
    from docx import Document
    doc = Document(io.BytesIO(content))
    all_records = []
    for table in doc.tables:
        if len(table.rows) < 2: continue
        headers = [normalize_col(cell.text) for cell in table.rows[0].cells]
        for row in table.rows[1:]:
            rec = {}
            for i, cell in enumerate(row.cells):
                if i < len(headers) and headers[i]:
                    rec[headers[i]] = clean_val(cell.text)
            if rec.get("policyno"): all_records.append(rec)
    if not all_records: return 0, 0, "No recognizable policy data in document"
    ins, upd = upsert_records(all_records)
    return ins, upd, None

# ── Endpoints ───────────────────────────────────────────────────────────────────

@app.post("/upload")
async def upload(files: list[UploadFile] = File(...)):
    results = {"uploaded": 0, "inserted": 0, "updated": 0, "saved_copies": [], "errors": []}
    for f in files:
        try:
            content = await f.read()
            # save a backup copy of every uploaded file
            saved_path = save_upload_copy(f.filename, content)
            results["saved_copies"].append(os.path.basename(saved_path))

            ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
            if ext in ("xlsx", "xls", "xlsm", "csv"):
                ins, upd, err = process_spreadsheet(content, f.filename)
            elif ext in ("jpg", "jpeg", "png"):
                ins, upd, err = process_image(content, f.filename)
            elif ext in ("doc", "docx"):
                ins, upd, err = process_docx(content, f.filename)
            else:
                results["errors"].append({"file": f.filename, "error": f"Unsupported format: .{ext}"})
                continue
            if err:
                results["errors"].append({"file": f.filename, "error": err})
            results["uploaded"] += 1
            results["inserted"] += ins
            results["updated"] += upd
        except Exception as e:
            results["errors"].append({"file": f.filename, "error": str(e)})
    db_push()  # sync to Turso cloud
    return results

@app.get("/search")
def search_policy(q: str = Query(..., min_length=1), field: str = Query("policyno")):
    """Search policies by policy number or name (partial match)."""
    query_val = q.strip().upper()
    with get_db() as conn:
        if field == "name":
            rows = conn.execute(
                "SELECT * FROM policies WHERE UPPER(COALESCE(name,'')) LIKE ? ORDER BY name, policyno",
                (f"%{query_val}%",)
            ).fetchall()
        elif field == "both":
            rows = conn.execute(
                "SELECT * FROM policies WHERE UPPER(TRIM(policyno)) LIKE ? OR UPPER(COALESCE(name,'')) LIKE ? ORDER BY policyno",
                (f"%{query_val}%", f"%{query_val}%")
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM policies WHERE UPPER(TRIM(policyno)) LIKE ? ORDER BY policyno",
                (f"%{query_val}%",)
            ).fetchall()
    return [dict(r) for r in rows]

@app.get("/master")
def get_master(limit: int = Query(50, ge=1, le=1000), offset: int = Query(0, ge=0)):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM policies ORDER BY id DESC LIMIT ? OFFSET ?", (limit, offset)).fetchall()
    return [dict(r) for r in rows]

@app.get("/master/count")
def get_count():
    with get_db() as conn:
        count = conn.execute("SELECT COUNT(*) AS cnt FROM policies").fetchone()["cnt"]
        last = conn.execute("SELECT MAX(updated_at) AS lu FROM policies").fetchone()["lu"]
    return {"count": count, "last_updated": last}

@app.delete("/master")
def clear_master(confirm: str = Query(...)):
    if confirm != "yes": raise HTTPException(400, "Pass ?confirm=yes to clear")
    with get_db() as conn:
        conn.execute("DELETE FROM policies")
    db_push()  # sync to Turso cloud
    return {"message": "All records deleted"}

@app.get("/uploads/history")
def get_upload_history(limit: int = Query(200, ge=1, le=1000)):
    """Return all uploaded data grouped by file/sheet, latest first."""
    from fastapi.responses import JSONResponse
    import math

    with get_db() as conn:
        batches = conn.execute(
            "SELECT id, filename, sheet_name, uploaded_at, record_count, records_json "
            "FROM upload_batches ORDER BY uploaded_at DESC, id DESC LIMIT ?",
            (limit,)
        ).fetchall()
    result = []
    for b in batches:
        entry = dict(b)
        try:
            records = json.loads(entry.pop("records_json") or "[]")
            entry["records"] = _scrub_nans(records)
        except (json.JSONDecodeError, TypeError):
            entry.pop("records_json", None)
            entry["records"] = []
        result.append(entry)

    # Use a NaN-safe serializer to prevent crashes on dirty legacy data
    def nan_safe_default(obj):
        if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
            return None
        raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")

    content = json.loads(json.dumps(result, default=nan_safe_default, allow_nan=False))
    return JSONResponse(content=content)

@app.delete("/uploads/history/{batch_id}")
def delete_upload_batch(batch_id: int):
    """Delete a single upload batch by ID."""
    with get_db() as conn:
        row = conn.execute("SELECT id FROM upload_batches WHERE id=?", (batch_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Batch not found")
        conn.execute("DELETE FROM upload_batches WHERE id=?", (batch_id,))
    db_push()
    return {"message": "Batch deleted", "id": batch_id}

@app.post("/generate-output")
async def generate_output(file: UploadFile = File(...)):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    # find header row (scan first 5 rows)
    header_row, pno_col, col_map = None, None, {}
    for r in range(1, min(6, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            norm = normalize_col(val) if val else None
            if norm == "policyno":
                header_row, pno_col = r, c
                break
        if header_row: break
    if not header_row: raise HTTPException(400, "No Policy No column found in authority file")

    # map all columns in header row
    for c in range(1, ws.max_column + 1):
        norm = normalize_col(ws.cell(header_row, c).value)
        if norm: col_map[norm] = c

    # ensure mobileno and status columns exist
    next_col = ws.max_column + 1
    if "mobileno" not in col_map:
        ws.cell(header_row, next_col, "Mobile No")
        col_map["mobileno"] = next_col
        next_col += 1
    if "status" not in col_map:
        ws.cell(header_row, next_col, "Status")
        col_map["status"] = next_col
        next_col += 1

    # load all DB records into dict keyed by uppercase policyno
    with get_db() as conn:
        all_db = conn.execute("SELECT * FROM policies").fetchall()
    db_map = {row["policyno"].upper().strip(): dict(row) for row in all_db if row["policyno"]}

    enriched, mobiles_added, data_start = 0, 0, header_row + 1

    for r in range(data_start, ws.max_row + 1):
        raw_pno = ws.cell(r, pno_col).value
        pno = normalize_policyno(raw_pno)
        if not pno: continue
        db_rec = db_map.get(pno)
        if not db_rec: continue

        # fill empty cells from DB (never overwrite non-empty)
        filled = False
        for field, col_idx in col_map.items():
            if field in ("mobileno", "status"): continue
            cell = ws.cell(r, col_idx)
            if cell.value is None or str(cell.value).strip() == "":
                db_val = db_rec.get(field)
                if db_val:
                    cell.value = db_val
                    filled = True

        # always fill mobileno
        mob = db_rec.get("mobileno")
        if mob:
            ws.cell(r, col_map["mobileno"], mob)
            mobiles_added += 1

        # status: fill only autodebit/dailycollection/branchpaid
        st = normalize_status(db_rec.get("status"))
        if st in ("autodebit", "dailycollection", "branchpaid"):
            ws.cell(r, col_map["status"], st)
            filled = True

        if filled: enriched += 1

    # auto-fit column widths
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, col_idx).value
            if val: max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"output_{date.today().isoformat()}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "X-Enriched": str(enriched),
            "X-Mobiles-Added": str(mobiles_added),
        }
    )

# serve index.html at root
@app.get("/")
def root():
    html_path = os.path.join(BASE_DIR, "index.html")
    return HTMLResponse(open(html_path, encoding="utf-8").read())

@app.api_route("/health", methods=["GET", "HEAD"])
def health():
    """Health check endpoint — pinged by UptimeRobot to prevent Render sleep."""
    return {"status": "ok"}
